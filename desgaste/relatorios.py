from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def gerar_excel_ocorrencias_inspecao(ocorrencias):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ocorrências"

    headers = [
        "Data da inspeção",
        "Hora início",
        "Hora fim",
        "Setor",
        "Via",
        "Trilho",
        "MT do problema",
        "Item",
        "Criticidade",
        "Responsável",
        "Foto",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="6F2DBD")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for ocorrencia in ocorrencias:
        inspecao = ocorrencia.inspecao

        foto_url = ""
        if ocorrencia.foto:
            try:
                foto_url = ocorrencia.foto.url
            except Exception:
                foto_url = ""

        ws.append([
            inspecao.data_inspecao.strftime("%d/%m/%Y") if inspecao.data_inspecao else "",
            inspecao.hora_inspecao.strftime("%H:%M") if inspecao.hora_inspecao else "",
            inspecao.hora_fim_inspecao.strftime("%H:%M") if inspecao.hora_fim_inspecao else "",
            inspecao.setor.nome if inspecao.setor else "",
            inspecao.get_via_display() if inspecao.via else "",
            ocorrencia.get_trilho_display() if ocorrencia.trilho else "",
            ocorrencia.mt_problema or "",
            ocorrencia.item.nome if ocorrencia.item else "",
            ocorrencia.get_criticidade_display() if ocorrencia.criticidade else "",
            inspecao.responsavel or "",
            foto_url,
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        1: 18,
        2: 14,
        3: 14,
        4: 28,
        5: 14,
        6: 14,
        7: 18,
        8: 36,
        9: 16,
        10: 24,
        11: 45,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 24

    return wb