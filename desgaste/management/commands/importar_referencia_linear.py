from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from desgaste.models import ReferenciaLinearPKMT


class Command(BaseCommand):
    help = "Importa a planilha de referência linear MT/PK para o banco."

    def add_arguments(self, parser):
        parser.add_argument(
            "arquivo",
            type=str,
            help="Caminho do arquivo Excel (.xlsx)"
        )
        parser.add_argument(
            "--limpar",
            action="store_true",
            help="Apaga os registros atuais antes de importar."
        )

    def handle(self, *args, **options):
        caminho_arquivo = Path(options["arquivo"])

        if not caminho_arquivo.exists():
            raise CommandError(f"Arquivo não encontrado: {caminho_arquivo}")

        if options["limpar"]:
            total_apagados = ReferenciaLinearPKMT.objects.count()
            ReferenciaLinearPKMT.objects.all().delete()
            self.stdout.write(
                self.style.WARNING(f"{total_apagados} registros antigos removidos.")
            )

        try:
            workbook = load_workbook(caminho_arquivo, data_only=True)
        except Exception as e:
            raise CommandError(f"Erro ao abrir o Excel: {e}")

        mapeamento_abas = {
            "Via 1": "1",
            "Via 2": "2",
        }

        total_importados = 0

        for nome_aba, via in mapeamento_abas.items():
            if nome_aba not in workbook.sheetnames:
                self.stdout.write(
                    self.style.WARNING(f"Aba '{nome_aba}' não encontrada. Pulando.")
                )
                continue

            ws = workbook[nome_aba]

            for row in ws.iter_rows(min_row=2, values_only=True):
                mt, pk_excel, local_excel = row

                if mt in (None, "") and pk_excel in (None, "") and local_excel in (None, ""):
                    continue

                mt_str = str(mt).strip() if mt is not None else ""
                pk_str = str(pk_excel).strip() if pk_excel is not None else ""
                local_str = str(local_excel).strip() if local_excel is not None else ""

                # Ignora linhas completamente inválidas
                if not local_str:
                    continue

                # Ajuste da nomenclatura do Excel para a operação
                if local_str.upper() == "PATIO1":
                    local_str = "PCR"
                elif local_str.upper() == "PATIO2":
                    local_str = "PCR"

                ReferenciaLinearPKMT.objects.create(
                    via=via,
                    marco_topografico=mt_str,
                    ponto_kilometrico=pk_str,
                    local_excel=local_str,
                )
                total_importados += 1

        self.stdout.write(
            self.style.SUCCESS(f"Importação concluída com {total_importados} registros.")
        )